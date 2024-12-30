import os
import pandas as pd
import json
import time
import random
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows  # 添加这一行
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import traceback

# 配置 ChromeOptions
options = Options()
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--headless')  # 启动无头模式
options.add_argument('--disable-gpu')
options.add_argument('--start-maximized')
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")

# 显式配置 Service，指定 ChromeDriver 路径
service = Service(executable_path='/usr/local/bin/chromedriver')  # 请根据实际路径修改

# 初始化 WebDriver
driver = webdriver.Chrome(service=service, options=options)

# 文件保存路径
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
html_output_path = os.path.join(desktop_path, "html_table_data20241225.xlsx")
json_output_path = os.path.join(desktop_path, "json_data20241224.json")

# 初始化 HTML 表格数据存储
all_html_data = pd.DataFrame()

# 初始化 JSON 数据存储
all_json_data = {}

def process_page(driver, url):
    try:
        # 等待页面加载
        WebDriverWait(driver, 30).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        
        # 获取表格数据
        selectors = [
            "#tabs-leftTabs--tabpanel-1 table tbody tr td.g-table-cell a",
            "table tbody tr td.g-table-cell.g-table-cell-fix-left a",
            "a.css-1qtqy4u"
        ]
        
        table_links = []
        for selector in selectors:
            print(f"尝试选择器: {selector}")  # 添加调试信息
            links = driver.find_elements(By.CSS_SELECTOR, selector)
            if links:
                table_links = links
                print(f"使用选择器 '{selector}' 找到 {len(links)} 个链接")
                break
        
        if not table_links:
            print(f"在 {url} 未找到任何链接")
            return None, None

        # 处理表格数据
        table = driver.find_element(By.XPATH, '//table')
        if not table:
            print(f"在 {url} 未找到表格")
            return None, None
            
        table_html = table.get_attribute('outerHTML')
        soup = BeautifulSoup(table_html, 'html.parser')
        
        # 提取表头和数据
        headers = [header.text.strip() for header in soup.find_all('th')]
        rows = []
        for row in soup.find_all('tr')[1:]:  # 跳过表头行
            cols = row.find_all('td')
            if cols:
                row_data = [col.text.strip() for col in cols]
                rows.append(row_data)
                
        # 创建基础DataFrame
        df = pd.DataFrame(rows, columns=headers)
        
        # 处理链接数据
        processed_tokens = set()
        token_data = {
            'Token Address': [],
            'Token Name': [],
            'Token Link': []
        }
        
        for link in table_links:
            try:
                href = link.get_attribute('href')
                text = link.text.strip()
                
                if 'token/' in href:
                    token_address = href.split('token/')[-1]
                    if token_address not in processed_tokens:
                        processed_tokens.add(token_address)
                        token_data['Token Address'].append(token_address)
                        token_data['Token Name'].append(text)
                        token_data['Token Link'].append(href)
                elif 'search?q=$' in href:  # 添加对搜索链接的处理
                    token_name = href.split('search?q=$')[-1]
                    if token_name not in processed_tokens:
                        processed_tokens.add(token_name)
                        token_data['Token Name'].append(token_name)
                        token_data['Token Address'].append('')  # 空地址
                        token_data['Token Link'].append(href)
                        
            except Exception as e:
                print(f"处理链接时出错: {str(e)}")
                continue
                
        # 创建链接数据DataFrame
        link_df = pd.DataFrame(token_data)
        
        # 合并数据
        if not df.empty and not link_df.empty:
            min_rows = min(len(df), len(link_df))
            df = df.iloc[:min_rows]
            link_df = link_df.iloc[:min_rows]
            df = pd.concat([df, link_df], axis=1)
            
        # 添加来源URL和时间戳
        df['Source URL'] = url
        df['Timestamp'] = pd.Timestamp.now()
        
        # 获取JSON数据
        script_tag = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//script[@id="__NEXT_DATA__"]'))
        )
        json_data = json.loads(script_tag.get_attribute('innerHTML'))
        
        return df, json_data
        
    except Exception as e:
        print(f"处理页面时出错: {str(e)}")
        traceback.print_exc()
        return None, None

def process_address(address, retry_count=3):
    """处理单个地址的数据采集"""
    url = f"https://gmgn.ai/sol/address/{address}"
    
    for attempt in range(retry_count):
        driver = None
        try:
            print(f"\n尝试第 {attempt + 1} 次处理地址: {address}")
            driver = create_new_driver()
            driver.set_page_load_timeout(30)
            
            # 访问页面
            driver.get(url)
            print(f"正在加载页面: {url}")
            
            # 等待页面加载完成
            WebDriverWait(driver, 60).until(
                lambda d: d.execute_script('return document.readyState') == 'complete'
            )
            
            # 等待表格容器加载
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '#tabs-leftTabs--tabpanel-1'))
            )
            
            # 获取表格数据
            table_data = None
            link_data = None
            
            # 1. 处理表格数据
            table = driver.find_element(By.XPATH, '//table')
            if table:
                table_html = table.get_attribute('outerHTML')
                soup = BeautifulSoup(table_html, 'html.parser')
                headers = [header.text.strip() for header in soup.find_all('th')]
                
                rows = soup.find_all('tr')
                data = []
                for row in rows:
                    cols = row.find_all('td')
                    if cols:
                        row_data = [col.text.strip() for col in cols]
                        data.append(row_data)
                
                table_data = pd.DataFrame(data, columns=headers)
                table_data["Source URL"] = url
            
            # 2. 处理链接数据
            selectors = [
                "#tabs-leftTabs--tabpanel-1 table tbody tr td.g-table-cell a",
                "table tbody tr td.g-table-cell.g-table-cell-fix-left a",
                "a.css-1qtqy4u"
            ]
            
            for selector in selectors:
                print(f"尝试选择器: {selector}")
                links = driver.find_elements(By.CSS_SELECTOR, selector)
                if links:
                    processed_tokens = set()
                    token_names = []
                    token_addresses = []
                    
                    for link in links:
                        try:
                            href = link.get_attribute('href')
                            text = link.text.strip()
                            
                            if 'token/' in href:
                                token_address = href.split('token/')[-1]
                                if token_address not in processed_tokens:
                                    processed_tokens.add(token_address)
                                    token_addresses.append(token_address)
                                    token_names.append(text)
                            elif 'search?q=$' in href:
                                token_name = href.split('search?q=$')[-1]
                                token_names.append(token_name)
                        except Exception as link_error:
                            print(f"处理链接时出错: {str(link_error)}")
                            continue
                    
                    # 创建链接数据的DataFrame
                    if token_addresses or token_names:
                        max_length = max(len(token_addresses), len(token_names))
                        token_addresses = token_addresses + [''] * (max_length - len(token_addresses))
                        token_names = token_names + [''] * (max_length - len(token_names))
                        
                        link_data = pd.DataFrame({
                            'Token Address': token_addresses,
                            'Token Name': token_names,
                            'Source URL': [url] * max_length
                        })
                    break
            
            # 获取 JSON 数据
            script_tag = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//script[@id="__NEXT_DATA__"]'))
            )
            
            json_data = None
            if script_tag:
                json_data = json.loads(script_tag.get_attribute('innerHTML'))
            
            return True, json_data, table_data, link_data
                
        except Exception as e:
            print(f"处理地址 {address} 时出错: {str(e)}")
            if attempt < retry_count - 1:
                delay = random.uniform(10, 15)
                print(f"等待 {delay:.2f} 秒后重试...")
                time.sleep(delay)
            else:
                print(f"已达到最大重试次数，跳过地址: {address}")
                return False, None, None, None
                
        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass
            
            if attempt < retry_count - 1:
                time.sleep(random.uniform(5, 8))

# 主程序部分
def main():
    try:
        # 获取用户输入
        print("请输入地址后缀（以逗号分隔）：")
        input_suffix = input().strip()
        address_list = [item.strip() for item in input_suffix.split(',') if item.strip()]
        
        if not address_list:
            print("没有输入有效的地址后缀！")
            return
            
        all_html_data = pd.DataFrame()
        all_json_data = {}
        
        for suffix in address_list:
            url = f"https://gmgn.ai/sol/address/{suffix}"
            print(f"\n正在处理: {url}")
            
            try:
                driver.get(url)
                df, json_data = process_page(driver, url)
                
                if df is not None:
                    all_html_data = pd.concat([all_html_data, df], ignore_index=True)
                if json_data is not None:
                    all_json_data[url] = json_data
                    
            except Exception as e:
                print(f"处理地址 {suffix} 时出错: {str(e)}")
                continue
                
            # 添加随机延迟
            delay = random.uniform(5, 10)
            print(f"等待 {delay:.2f} 秒...")
            time.sleep(delay)
            
        # 保存数据
        if not all_html_data.empty:
            save_to_excel(all_html_data, html_output_path)
        if all_json_data:
            save_to_json(all_json_data, json_output_path)
            
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
