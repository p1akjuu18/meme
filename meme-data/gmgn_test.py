import os
import pandas as pd
import json
import time
import random
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置 ChromeOptions
options = Options()
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--headless')  # 启动无头模式
options.add_argument('--disable-gpu')
options.add_argument('--start-maximized')
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")

# 显式配置 Service，指定 ChromeDriver 路径
service = Service(executable_path='/usr/local/bin/chromedriver')  # 请根据实际路径修改

# 初始化 WebDriver
driver = webdriver.Chrome(service=service, options=options)

# 文件保存路径
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
html_output_path = os.path.join(desktop_path, "html_table_data20241222.xlsx")
json_output_path = os.path.join(desktop_path, "json_data20241222.json")

# 初始化 HTML 表格数据存储
all_html_data = pd.DataFrame()

# 初始化 JSON 数据存储
all_json_data = {}

# 手动输入地址列表
print("请输入地址后缀（以逗号分隔）：")
input_suffix = input().strip()  # 获取用户输入
address_list = [item.strip() for item in input_suffix.split(',') if item.strip()]  # 处理输入，去除空白

# 检查地址列表是否为空
if not address_list:
    print("没有输入有效的地址后缀，程序结束！")
else:
    # 遍历每个地址，进行爬取
    for suffix in address_list:
        url = f"https://gmgn.ai/sol/address/{suffix}"
        print(f"正在爬取：{url}")

        try:
            driver.get(url)

            # 等待 JSON 数据加载完成
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//script[@id="__NEXT_DATA__"]'))
            )

            # 提取 JSON 数据
            script_tag = driver.find_element(By.XPATH, '//script[@id="__NEXT_DATA__"]')
            if script_tag:
                json_data = json.loads(script_tag.get_attribute('innerHTML'))
                all_json_data[url] = json_data  # 保存 JSON 数据
            else:
                print(f"未能在 {url} 找到 JSON 数据")
                continue

            # HTML 数据爬取部分开始
            # 1. 爬取 css-1qtqy4u 链接
            token_links = WebDriverWait(driver, 30).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'a.css-1qtqy4u'))
            )
            
            # 提取链接信息
            token_data = []
            for link in token_links:
                token_info = {
                    'text': link.text,
                    'href': link.get_attribute('href'),
                }
                token_data.append(token_info)
            
            # 将 token 数据添加到 DataFrame
            for token in token_data:
                df_token = pd.DataFrame({
                    'Token Name': [token['text']],
                    'Token URL': [token['href']],
                    'Source URL': [url]
                })
                all_html_data = pd.concat([all_html_data, df_token], ignore_index=True)

            # 2. 爬取表格数据
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//table'))
            )

            # 提取表格 HTML 数据
            table = driver.find_element(By.XPATH, '//table')
            if table:
                table_html = table.get_attribute('outerHTML')

                # **保存表格 HTML 到本地文件**
                with open(f"{desktop_path}/table_{suffix}.html", "w", encoding="utf-8") as f:
                    f.write(table_html)
                print(f"已保存表格 HTML 到文件：table_{suffix}.html")

                # 使用 BeautifulSoup 解析表格
                soup = BeautifulSoup(table_html, 'html.parser')

                # 提取表头
                headers = [header.text.strip() for header in soup.find_all('th')]

                # 提取表格行内容
                rows = soup.find_all('tr')
                data = []
                for row in rows:
                    cols = row.find_all('td')
                    if cols:
                        data.append([col.text.strip() for col in cols])

                # 转换为 DataFrame
                df = pd.DataFrame(data, columns=headers) if headers else pd.DataFrame(data)

                # 添加来源 URL 列
                df["Source URL"] = url

                # 打印当前爬取结果
                print(f"爬取结果（HTML 表格）：\n{df.head()}")

                # 将当前数据追加到总数据中
                all_html_data = pd.concat([all_html_data, df], ignore_index=True)

            else:
                print(f"未能在 {url} 找到表格数据")
                continue

        except Exception as e:
            print(f"爬取 {url} 时发生错误：{e}")

        # 随机休息 5 到 10 秒（可根据需要调整时间）
        time.sleep(random.uniform(5, 10))

    # 处理 Excel 文件追加
    if os.path.exists(html_output_path):
        # 如果文件已经存在，则加载现有的工作簿
        wb = load_workbook(html_output_path)
        sheet = wb['Data'] if 'Data' in wb.sheetnames else wb.create_sheet('Data')

        # 获取现有数据的最后一行
        startrow = sheet.max_row

        # 将新数据添加到 Excel 文件
        for r in dataframe_to_rows(all_html_data, index=False, header=False):
            sheet.append(r)

        # 保存更新后的文件
        wb.save(html_output_path)
    else:
        # 如果文件不存在，直接创建一个新的 Excel 文件
        all_html_data.to_excel(html_output_path, index=False, sheet_name='Data')

    print(f"HTML 表格数据已追加到：{html_output_path}")

    # 读取现有的 JSON 数据并追加新数据
    if os.path.exists(json_output_path):
        with open(json_output_path, "r", encoding="utf-8") as f:
            existing_json_data = json.load(f)
        existing_json_data.update(all_json_data)  # 合并新的 JSON 数据
    else:
        existing_json_data = all_json_data

    # 保存 JSON 数据到 JSON 文件
    with open(json_output_path, "w", encoding="utf-8") as f:
        json.dump(existing_json_data, f, indent=4, ensure_ascii=False)
    print(f"JSON 数据已追加到：{json_output_path}")

# 关闭浏览器
driver.quit()
